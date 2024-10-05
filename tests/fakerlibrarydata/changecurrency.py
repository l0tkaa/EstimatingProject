import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NumberFormatDescriptor

# Folder path where Excel files are located
folder_path = r'C:\programming projects\sample folder\excel samples'

# Loop through all Excel files in the directory
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(folder_path, filename)
        
        # Load the Excel file into a DataFrame
        df = pd.read_excel(file_path)

        # Display original data (optional)
        print(f"Original Data in {filename}:")

        # Save the DataFrame back to Excel to apply formatting
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, index=False)

            # Access the workbook and the active sheet
            workbook = writer.book
            sheet_names = writer.sheets.keys() #Get all sheet names
            first_sheet_name = list(sheet_names)[0] #Convert to list and get the first sheet name
            worksheet = workbook[first_sheet_name] #access the first sheet

            # loop through all columns in the DataFrame to find "Price" columns
            for col in worksheet.iter_cols(min_row=1, max_row=1): # Check the header row
                for cell in col:
                    if "Price" in str(cell.value):
                        # Apply currency format to all cells in the column
                        for row in range(2, worksheet.max_row +1): # Skip header row
                            worksheet.cell(row=row, column=cell.col_idx).number_format = '$#,##0.00'

        print(f"Updated currency format in {filename}")